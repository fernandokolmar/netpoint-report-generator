"""
Gerador de arquivos Excel para relatórios PRSA.

Este módulo contém a classe ExcelGenerator que cria arquivos Excel
com múltiplas planilhas, tabelas, fórmulas e gráficos.
"""

from typing import Dict, Optional, Callable
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.text import RichText
from openpyxl.chart.title import Title
from openpyxl.drawing.text import (
    RichTextProperties,
    Paragraph,
    ParagraphProperties,
    CharacterProperties,
    RegularTextRun
)
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties

from config import settings


class ExcelGenerator:
    """
    Gera arquivos Excel com relatórios formatados.

    Esta classe é responsável por:
    - Criar workbook Excel com múltiplas planilhas
    - Criar tabelas Excel com formatação
    - Adicionar fórmulas Excel (XLOOKUP, SUBTOTAL, MAX)
    - Criar gráficos de linha
    - Ajustar larguras de colunas automaticamente

    Attributes:
        progress_callback: Função opcional para reportar progresso
    """

    def __init__(self, progress_callback: Optional[Callable[[str], None]] = None):
        """
        Inicializa o gerador de Excel.

        Args:
            progress_callback: Função opcional que recebe mensagens de progresso
        """
        self.progress_callback = progress_callback

    def _log(self, message: str) -> None:
        """
        Envia mensagem de log via callback se disponível.

        Args:
            message: Mensagem a ser logada
        """
        if self.progress_callback:
            self.progress_callback(message)

    def _sanitize_column_name(self, name: str) -> str:
        """
        Sanitiza nome de coluna para evitar problemas no XML do Excel.

        Remove ou substitui caracteres problemáticos como parênteses,
        colchetes, e outros caracteres especiais.

        Args:
            name: Nome original da coluna

        Returns:
            Nome sanitizado compatível com Excel
        """
        if not name:
            return "Coluna"

        # Converter para string se necessário
        name = str(name)

        # Substituir caracteres problemáticos
        replacements = {
            '(': '_',
            ')': '_',
            '[': '_',
            ']': '_',
            ':': '_',
            '/': '_',
            '\\': '_',
            '*': '_',
            '?': '_',
            '"': '_',
            '<': '_',
            '>': '_',
            '|': '_',
            '#': '_',
            "'": '_',
        }

        for char, replacement in replacements.items():
            name = name.replace(char, replacement)

        # Remover underscores duplos ou no final
        while '__' in name:
            name = name.replace('__', '_')
        name = name.strip('_')

        return name if name else "Coluna"

    def _ensure_unique_column_names(self, columns: list) -> list:
        """
        Garante que todos os nomes de colunas sao unicos.

        O Excel corrompe tabelas que tem colunas com nomes duplicados.
        Esta funcao adiciona sufixos numericos para colunas duplicadas.

        Args:
            columns: Lista de nomes de colunas (ja sanitizados)

        Returns:
            Lista de nomes de colunas unicos
        """
        seen = {}
        unique_columns = []

        for col in columns:
            if col in seen:
                # Coluna duplicada - adicionar sufixo numerico
                seen[col] += 1
                unique_col = f"{col}_{seen[col]}"
                # Garantir que o novo nome tambem e unico
                while unique_col in seen:
                    seen[col] += 1
                    unique_col = f"{col}_{seen[col]}"
                unique_columns.append(unique_col)
                seen[unique_col] = 0
            else:
                seen[col] = 0
                unique_columns.append(col)

        return unique_columns

    def _convert_cell_value(self, value):
        """
        Converte valor de célula para tipo compatível com Excel.

        Trata valores NaN, tipos numpy, e outros tipos problemáticos.

        Args:
            value: Valor a ser convertido

        Returns:
            Valor convertido (str, int, float, bool, None ou datetime)
        """
        # Verificar NaN/None primeiro
        try:
            if value is None:
                return None
            if pd.isna(value):
                return None
        except (ValueError, TypeError):
            pass

        # Converter tipos numpy para tipos Python nativos
        if hasattr(value, 'item'):
            try:
                return value.item()
            except (ValueError, TypeError):
                return str(value)

        # Verificar se é um tipo básico aceito pelo Excel
        if isinstance(value, (str, int, float, bool)):
            return value

        # Converter qualquer outro tipo para string
        try:
            return str(value)
        except:
            return None

    def generate(self, dfs: Dict[str, pd.DataFrame], output_path: str) -> str:
        """
        Gera arquivo Excel com todas as planilhas e formatações.

        Args:
            dfs: Dicionário com DataFrames processados
                Chaves esperadas: 'totalizado_processed', 'mensagens_processed',
                                 'relatorio_processed', 'inscritos_processed'
            output_path: Caminho completo onde salvar o arquivo Excel

        Returns:
            Caminho do arquivo gerado

        Raises:
            KeyError: Se alguma chave esperada estiver faltando no dicionário dfs
            IOError: Se houver erro ao salvar o arquivo
            ValueError: Se os DataFrames estiverem vazios ou inválidos

        Example:
            >>> generator = ExcelGenerator()
            >>> dfs = {
            ...     'totalizado_processed': df_totalizado,
            ...     'mensagens_processed': df_mensagens,
            ...     'relatorio_processed': df_relatorio,
            ...     'inscritos_processed': df_inscritos
            ... }
            >>> path = generator.generate(dfs, 'relatorio.xlsx')
            >>> print(f"Arquivo gerado: {path}")
        """
        self._log("Criando arquivo Excel...")

        # Resetar contador de IDs de tabela para cada novo arquivo
        ExcelGenerator._table_id_counter = 0

        # Flags para controle de abas opcionais
        has_mensagens = 'mensagens_processed' in dfs and dfs['mensagens_processed'] is not None and len(dfs['mensagens_processed']) > 0
        has_chat = 'chat_processed' in dfs and dfs['chat_processed'] is not None and len(dfs['chat_processed']) > 0

        # Coletar enquetes disponíveis (ordenadas)
        enquete_keys = sorted([k for k in dfs if k.startswith('enquete_') and k.endswith('_processed')])
        enquetes = [(k, dfs[k]) for k in enquete_keys if dfs[k] is not None and len(dfs[k]) > 0]

        # Criar workbook
        wb = Workbook()

        # Remover a planilha padrão
        wb.remove(wb.active)

        # Criar planilhas na ordem correta
        self._log("Criando planilha Retenção...")
        self._create_retencao_sheet(wb, dfs['totalizado_processed'], has_mensagens, has_chat)

        # Criar planilha Mensagens (se houver dados)
        if has_mensagens:
            self._log("Criando planilha Mensagens...")
            self._create_mensagens_sheet(wb, dfs['mensagens_processed'])
        else:
            self._log("⊘ Planilha Mensagens: não há dados")

        # Criar planilha Chat (se houver dados)
        if has_chat:
            self._log("Criando planilha Chat...")
            self._create_chat_sheet(wb, dfs['chat_processed'])
        else:
            self._log("⊘ Planilha Chat: não há dados")

        self._log("Criando planilha Acessos...")
        self._create_acessos_sheet(wb, dfs['relatorio_processed'])

        self._log("Criando planilha Inscritos...")
        self._create_inscritos_sheet(wb, dfs['inscritos_processed'])

        # Criar planilhas de enquete (se houver)
        for i, (key, df_enquete) in enumerate(enquetes, start=1):
            sheet_name = f"Enquete {i:02d}"
            self._log(f"Criando planilha {sheet_name}...")
            self._create_enquete_sheet(wb, df_enquete, sheet_name, i)

        if not enquetes:
            self._log("⊘ Nenhuma enquete para incluir no relatório")

        # Finalizar fórmulas cross-sheet (DEPOIS que todas as tabelas existem)
        self._log("Finalizando fórmulas da tabela Resumo...")
        self._finalize_resumo_formulas(wb, has_mensagens, has_chat)

        # Salvar arquivo
        self._log(f"Salvando arquivo: {output_path}")
        wb.save(output_path)

        self._log(f"Arquivo salvo com sucesso: {output_path}")
        return output_path

    def _create_retencao_sheet(self, wb: Workbook, df: pd.DataFrame, has_mensagens: bool = True, has_chat: bool = False) -> None:
        """
        Cria planilha "Retenção na live" com gráfico e resumo estatístico.

        Esta planilha contém:
        - Tabela com dados de retenção (Horário, Usuários conectados, Max)
        - Gráfico de linha mostrando usuários conectados por horário
        - Tabela resumo com estatísticas (inscritos, pico de audiência, etc.)

        Args:
            wb: Workbook do openpyxl
            df: DataFrame com dados de retenção (colunas: Data, Usuarios conectados)
            has_mensagens: Se True, inclui linha de mensagens no resumo
            has_chat: Se True, inclui linha de chat no resumo
        """
        ws = wb.create_sheet("Retencao na live")

        # Cabeçalhos da tabela principal (nova ordem: Horário, Usuários, Max)
        ws['A1'] = 'Horário'
        ws['B1'] = 'Usuários conectados'
        ws['C1'] = 'Max'

        # Adicionar dados
        for i, (idx, row) in enumerate(df.iterrows()):
            row_num = i + 2  # Linha 2 é a primeira linha de dados
            # Coluna A: Horário formatado como texto HH:MM (extraído do datetime)
            data_value = row[settings.COL_DATA]
            if pd.notna(data_value):
                # Converter para string formatada HH:MM
                if hasattr(data_value, 'strftime'):
                    ws[f'A{row_num}'] = data_value.strftime('%H:%M')
                else:
                    # Se já for string ou outro tipo, tentar converter
                    ws[f'A{row_num}'] = str(data_value)
            else:
                ws[f'A{row_num}'] = ''
            # Coluna B: Usuários conectados
            ws[f'B{row_num}'] = row[settings.COL_USUARIOS_CONECTADOS]

        # Última linha de dados
        last_data_row = len(df) + 1

        # Encontrar o índice do PRIMEIRO valor máximo (evitar duplicações)
        max_value = 0
        max_row_index = 0
        for i, (idx, row) in enumerate(df.iterrows()):
            usuarios = row[settings.COL_USUARIOS_CONECTADOS]
            if usuarios is not None and usuarios > max_value:
                max_value = usuarios
                max_row_index = i

        # Criar tabela SEM linha de totais (apenas dados) - agora apenas 3 colunas
        tab_ref = f"A1:C{last_data_row}"
        self._create_table(
            ws,
            table_name=settings.TABLE_RETENCAO,
            ref=tab_ref
        )

        # Adicionar valor Max apenas na linha do primeiro máximo
        for i in range(len(df)):
            row_num = i + 2
            if i == max_row_index:
                # Apenas o primeiro máximo recebe o valor
                ws[f'C{row_num}'] = max_value
            else:
                # Demais células ficam com N/A (para não aparecer no gráfico)
                ws[f'C{row_num}'] = '=NA()'

        # Linha de totais FORA da tabela (como células normais)
        total_row = last_data_row + 1
        ws[f'A{total_row}'] = 'Total'
        ws[f'A{total_row}'].font = Font(bold=True)
        ws[f'B{total_row}'] = f"=MAX({settings.TABLE_RETENCAO}[Usuários conectados])"
        ws[f'B{total_row}'].font = Font(bold=True)

        # Criar estrutura da tabela de resumo (SEM fórmulas cross-sheet ainda)
        self._create_resumo_structure(ws, len(df), has_mensagens, has_chat)

        # Criar gráfico de linha
        self._create_retencao_chart(ws, len(df))

        # Ajustar larguras das colunas (agora apenas 3 colunas de dados)
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['P'].width = 25
        ws.column_dimensions['Q'].width = 20

    def _create_resumo_structure(self, ws: Worksheet, data_rows: int, has_mensagens: bool = True, has_chat: bool = False) -> None:
        """
        Cria estrutura da tabela de resumo (apenas com fórmulas locais).

        Fórmulas cross-sheet serão adicionadas depois em _finalize_resumo_formulas().

        Args:
            ws: Worksheet onde criar a tabela
            data_rows: Número de linhas de dados da tabela principal
            has_mensagens: Se True, inclui linha de mensagens
            has_chat: Se True, inclui linha de chat
        """
        # IMPORTANTE: Headers devem ter valores válidos (não vazios)
        # Posição ajustada para acomodar tabela com 3 colunas (era Q, agora P)
        ws['P1'] = 'Estatística'
        ws['Q1'] = 'Valor'

        # Labels fixas na coluna P (linhas 2-6)
        ws['P2'] = 'Quantidade de Inscritos'
        ws['P3'] = 'Usuários distintos na live'
        ws['P4'] = 'Pico de audiência'
        ws['P5'] = 'Hora de pico'
        ws['P6'] = 'Tempo médio assitido (hh:mm)'

        # Determinar próxima linha disponível
        next_row = 7

        # Adicionar linha de mensagens (se houver)
        if has_mensagens:
            ws[f'P{next_row}'] = 'Total mensagens enviadas'
            next_row += 1

        # Adicionar linha de chat (se houver)
        if has_chat:
            ws[f'P{next_row}'] = 'Total mensagens no chat'
            next_row += 1

        # Se não houver mensagens nem chat, adicionar linha placeholder
        if not has_mensagens and not has_chat:
            # Tabela precisa ter pelo menos estas linhas
            next_row = 7

        # Calcular última linha da tabela (mínimo 6, máximo 8)
        last_row = max(next_row - 1, 6)

        # Fórmulas locais na coluna Q
        ws['Q4'] = f"=MAX({settings.TABLE_RETENCAO}[Usuários conectados])"
        # Hora de pico: busca a primeira ocorrência do valor máximo na coluna Horário
        ws['Q5'] = (
            f"=_xlfn.XLOOKUP(MAX({settings.TABLE_RETENCAO}[Usuários conectados]),"
            f"{settings.TABLE_RETENCAO}[Usuários conectados],"
            f"{settings.TABLE_RETENCAO}[Horário],\"-\",0,1)"
        )
        ws['Q5'].alignment = Alignment(horizontal='right')
        # Q2, Q3, Q6 e outras serão preenchidas em _finalize_resumo_formulas

        # Criar tabela de resumo SEM linha de totais
        self._create_table(
            ws,
            table_name=settings.TABLE_RESUMO,
            ref=f"P1:Q{last_row}"
        )

    def _finalize_resumo_formulas(self, wb: Workbook, has_mensagens: bool = True, has_chat: bool = False) -> None:
        """
        Adiciona fórmulas cross-sheet na tabela de resumo.

        Deve ser chamado DEPOIS que todas as outras tabelas já foram criadas.

        Args:
            wb: Workbook com todas as planilhas
            has_mensagens: Se True, adiciona fórmula para mensagens
            has_chat: Se True, adiciona fórmula para chat
        """
        ws = wb["Retencao na live"]

        # Usar nomes sanitizados nas fórmulas
        data_cadastro = self._sanitize_column_name('Data de Cadastro')
        nome = self._sanitize_column_name('Nome')
        tempo_minutos = self._sanitize_column_name('Tempo_Minutos')
        data = self._sanitize_column_name('Data')
        data_formatada = self._sanitize_column_name('DataFormatada')

        # Agora que todas as tabelas existem, adicionar as fórmulas cross-sheet
        # Quantidade de inscritos = contagem da tabela inscritos
        ws['Q2'] = f"=SUBTOTAL(103,{settings.TABLE_INSCRITOS}[{data_cadastro}])"
        # Usuários distintos = contagem da tabela acessos
        ws['Q3'] = f"=SUBTOTAL(103,{settings.TABLE_ACESSOS}[{nome}])"
        # Tempo médio assistido
        ws['Q6'] = f"=AVERAGE({settings.TABLE_ACESSOS}[{tempo_minutos}])/1440"
        ws['Q6'].number_format = '[h]:mm'

        # Próxima linha disponível para fórmulas opcionais
        next_row = 7

        # Total mensagens (se houver)
        if has_mensagens:
            ws[f'Q{next_row}'] = f"=SUBTOTAL(103,{settings.TABLE_MENSAGENS}[{data}])"
            next_row += 1

        # Total mensagens no chat (se houver)
        if has_chat:
            ws[f'Q{next_row}'] = f"=SUBTOTAL(103,{settings.TABLE_CHAT}[{data_formatada}])"
            next_row += 1

    def _create_retencao_chart(self, ws: Worksheet, data_rows: int) -> None:
        """
        Cria gráfico de linha para visualização de retenção.

        Usa LineChart ao invés de AreaChart para maior compatibilidade.
        O gráfico mostra a evolução de usuários conectados ao longo do tempo.

        Características do gráfico:
        - Tipo: Gráfico de linha
        - Título descritivo
        - Cor azul sólido (similar ao estilo da tabela TableStyleMedium12)
        - Eixos formatados
        - Marcador vermelho no ponto de pico com rótulo

        Args:
            ws: Worksheet onde criar o gráfico
            data_rows: Número de linhas de dados
        """
        # Usar LineChart que é mais confiável
        chart = LineChart()

        # Título simples
        chart.title = "Usuários conectados minuto a minuto"

        # Estilo base (usar estilo padrão do Excel que é seguro)
        chart.style = 13

        # Configurar eixo Y (Quantidade de usuários)
        chart.y_axis.delete = False  # Garantir que o eixo Y seja exibido
        chart.y_axis.tickLblPos = "nextTo"  # Labels ao lado do eixo
        chart.y_axis.numFmt = "#,##0"  # Formato numérico (sem decimais, com separador de milhar)

        # Título do eixo Y - definir texto primeiro, depois customizar
        chart.y_axis.title = "Qtd. Usuários"

        # Customizar o título Y para texto horizontal e posição dentro da área de plotagem
        y_title = chart.y_axis.title
        y_title.overlay = True  # Permite sobrepor a área de plotagem
        y_title.txPr = RichText(
            bodyPr=RichTextProperties(
                vert='horz',  # Texto horizontal (não vertical/rotacionado)
                anchor='ctr',
            ),
            p=[Paragraph(
                pPr=ParagraphProperties(defRPr=CharacterProperties(sz=1000, b=True)),
                endParaRPr=CharacterProperties(sz=1000, b=True)
            )]
        )
        # Posicionar título Y centralizado verticalmente, próximo à borda esquerda interna
        y_title.layout = Layout(
            manualLayout=ManualLayout(
                xMode="edge",
                yMode="edge",
                x=0.06,   # 6% da esquerda (próximo à borda interna esquerda)
                y=0.42,   # 42% do topo (um pouco mais para o meio do eixo)
            )
        )

        # Configurar eixo X (Horário) com labels espaçados
        # Título do eixo X - definir texto primeiro, depois customizar
        chart.x_axis.title = "Horário"

        # Customizar o título X para posição dentro da área de plotagem
        x_title = chart.x_axis.title
        x_title.overlay = True  # Permite sobrepor a área de plotagem
        x_title.txPr = RichText(
            bodyPr=RichTextProperties(
                vert='horz',  # Texto horizontal
                anchor='ctr',
            ),
            p=[Paragraph(
                pPr=ParagraphProperties(defRPr=CharacterProperties(sz=1000, b=True)),
                endParaRPr=CharacterProperties(sz=1000, b=True)
            )]
        )
        # Posicionar título X centralizado horizontalmente, próximo à borda inferior interna
        x_title.layout = Layout(
            manualLayout=ManualLayout(
                xMode="edge",
                yMode="edge",
                x=0.50,   # 50% da largura (centralizado horizontalmente)
                y=0.82,   # 82% do topo (um pouco mais para cima)
            )
        )
        chart.x_axis.tickLblPos = "low"  # Labels embaixo do eixo
        chart.x_axis.tickLblSkip = 10  # Mostrar 1 label a cada 10
        chart.x_axis.tickMarkSkip = 10
        chart.x_axis.delete = False  # Garantir que o eixo X não seja deletado

        # Adicionar linhas de grade verticais (major gridlines) em cinza claro
        chart.x_axis.majorGridlines = ChartLines()
        chart.x_axis.majorGridlines.spPr = GraphicalProperties()
        chart.x_axis.majorGridlines.spPr.ln.solidFill = "D9D9D9"  # Cinza claro
        chart.x_axis.majorGridlines.spPr.ln.w = 9525  # Linha fina (0.75pt)

        # Dados para o gráfico - apenas valores numéricos (sem cabeçalho)
        # Coluna B = Usuários conectados (dados principais)
        data = Reference(ws, min_col=2, min_row=2, max_row=data_rows + 1)

        # Coluna A = Horário formatado (categorias para o eixo X)
        cats = Reference(ws, min_col=1, min_row=2, max_row=data_rows + 1)

        # Adicionar dados ao gráfico (sem títulos dos dados)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)

        # Deletar legenda (não é necessária pois só temos uma série)
        chart.legend = None

        # Encontrar índice do valor máximo para o marcador
        max_value = 0
        max_index = 0
        for i in range(data_rows):
            cell_value = ws.cell(row=i + 2, column=2).value
            if cell_value is not None and cell_value > max_value:
                max_value = cell_value
                max_index = i

        # Configurar a série para ter linha suave e cor azul sólido
        if chart.series:
            series = chart.series[0]
            series.smooth = True  # Linha suave
            # Aplicar cor azul sólido (similar ao TableStyleMedium12)
            series.graphicalProperties.line.solidFill = "4472C4"  # Azul Excel padrão
            series.graphicalProperties.line.width = 25000  # largura em EMUs (2.5pt)

            # Configurar marcadores - nenhum por padrão
            series.marker = Marker(symbol='none')

            # Habilitar rótulos de dados apenas para o ponto máximo
            series.labels = DataLabelList()
            series.labels.showVal = False  # Não mostrar valores por padrão
            series.labels.showCatName = False
            series.labels.showSerName = False

        # Criar segunda série apenas com o ponto máximo (para destaque vermelho)
        # Isso é feito usando a coluna C (Max) que só tem valor no ponto máximo
        max_data = Reference(ws, min_col=3, min_row=2, max_row=data_rows + 1)
        chart.add_data(max_data, titles_from_data=False)

        if len(chart.series) > 1:
            max_series = chart.series[1]
            # Sem linha, apenas marcador
            max_series.graphicalProperties.line.noFill = True
            # Marcador vermelho circular
            max_series.marker = Marker(symbol='circle', size=10)
            max_series.marker.graphicalProperties.solidFill = "FF0000"  # Vermelho
            max_series.marker.graphicalProperties.line.solidFill = "FF0000"

            # Mostrar rótulo de dados no ponto máximo
            max_series.labels = DataLabelList()
            max_series.labels.showVal = True
            max_series.labels.showCatName = True  # Mostrar horário
            max_series.labels.showSerName = False
            max_series.labels.separator = "\n"  # Quebra de linha entre horário e valor

        # Posicionar e dimensionar gráfico (tamanho grande)
        chart.width = 20  # largura em cm
        chart.height = 10  # altura em cm

        ws.add_chart(chart, "D2")

    def _create_mensagens_sheet(self, wb: Workbook, df: pd.DataFrame) -> None:
        """
        Cria planilha "Mensagens" com tabela de mensagens enviadas.

        Args:
            wb: Workbook do openpyxl
            df: DataFrame com mensagens
        """
        ws = wb.create_sheet("Mensagens")

        # Sanitizar nomes das colunas
        sanitized_list = [self._sanitize_column_name(col) for col in df.columns]

        # Garantir unicidade dos nomes (evita corrupcao do XML)
        sanitized_columns = self._ensure_unique_column_names(sanitized_list)

        # Escrever cabecalhos unicos
        for col_idx, col_name in enumerate(sanitized_columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Escrever dados usando índices numéricos para evitar erro de Series ambígua
        for row_idx in range(len(df)):
            for col_idx in range(len(df.columns)):
                value = df.iloc[row_idx, col_idx]
                # Converter tipos problemáticos para valores Python nativos
                value = self._convert_cell_value(value)
                ws.cell(row=row_idx + 2, column=col_idx + 1, value=value)

        # Última linha de dados
        last_data_row = len(df) + 1

        # Criar tabela SEM linha de totais
        tab_ref = f"A1:{get_column_letter(len(df.columns))}{last_data_row}"
        self._create_table(
            ws,
            table_name=settings.TABLE_MENSAGENS,
            ref=tab_ref
        )

        # Linha de totais FORA da tabela (como células normais)
        total_row = last_data_row + 1
        ws[f'A{total_row}'] = "Total"
        ws[f'A{total_row}'].font = Font(bold=True)
        last_col = len(df.columns)
        last_col_letter = get_column_letter(last_col)
        # Usar nome sanitizado da coluna Data
        data_col_name = self._sanitize_column_name('Data')
        ws[f'{last_col_letter}{total_row}'] = (
            f"=SUBTOTAL(103,{settings.TABLE_MENSAGENS}[{data_col_name}])"
        )
        ws[f'{last_col_letter}{total_row}'].font = Font(bold=True)

        # Ajustar larguras das colunas
        self._auto_adjust_column_widths(ws, max_width=50)

    def _create_chat_sheet(self, wb: Workbook, df: pd.DataFrame) -> None:
        """
        Cria planilha "Chat" com tabela de mensagens do chat.

        Remove as colunas 'Cliente' e 'Sala' que não são relevantes.

        Args:
            wb: Workbook do openpyxl
            df: DataFrame com mensagens do chat
        """
        ws = wb.create_sheet("Chat")

        # Remover colunas não desejadas (Cliente, Sala)
        columns_to_remove = settings.CHAT_COLUMNS_TO_REMOVE
        df_filtered = df.drop(columns=[col for col in columns_to_remove if col in df.columns], errors='ignore')

        # Sanitizar nomes das colunas
        sanitized_list = [self._sanitize_column_name(col) for col in df_filtered.columns]

        # Garantir unicidade dos nomes (evita corrupcao do XML)
        sanitized_columns = self._ensure_unique_column_names(sanitized_list)

        # Escrever cabecalhos unicos
        for col_idx, col_name in enumerate(sanitized_columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Escrever dados usando índices numéricos para evitar erro de Series ambígua
        for row_idx in range(len(df_filtered)):
            for col_idx in range(len(df_filtered.columns)):
                value = df_filtered.iloc[row_idx, col_idx]
                # Converter tipos problemáticos para valores Python nativos
                value = self._convert_cell_value(value)
                ws.cell(row=row_idx + 2, column=col_idx + 1, value=value)

        # Última linha de dados
        last_data_row = len(df_filtered) + 1

        # Criar tabela SEM linha de totais
        tab_ref = f"A1:{get_column_letter(len(df_filtered.columns))}{last_data_row}"
        self._create_table(
            ws,
            table_name=settings.TABLE_CHAT,
            ref=tab_ref
        )

        # Linha de totais FORA da tabela (como células normais)
        total_row = last_data_row + 1
        ws[f'A{total_row}'] = "Total"
        ws[f'A{total_row}'].font = Font(bold=True)
        last_col = len(df_filtered.columns)
        last_col_letter = get_column_letter(last_col)
        # Usar nome sanitizado da coluna DataFormatada
        data_col_name = self._sanitize_column_name('DataFormatada')
        ws[f'{last_col_letter}{total_row}'] = (
            f"=SUBTOTAL(103,{settings.TABLE_CHAT}[{data_col_name}])"
        )
        ws[f'{last_col_letter}{total_row}'].font = Font(bold=True)

        # Ajustar larguras das colunas
        self._auto_adjust_column_widths(ws, max_width=50)

    def _create_acessos_sheet(self, wb: Workbook, df: pd.DataFrame) -> None:
        """
        Cria planilha "Acessos" com relatório de acesso dos usuários.

        Args:
            wb: Workbook do openpyxl
            df: DataFrame com relatório de acesso
        """
        ws = wb.create_sheet("Acessos")

        # Sanitizar nomes das colunas
        sanitized_list = [self._sanitize_column_name(col) for col in df.columns]

        # Garantir unicidade dos nomes (evita corrupcao do XML)
        unique_columns = self._ensure_unique_column_names(sanitized_list)

        # Mapear nome original -> nome final unico
        sanitized_columns = {}
        for i, col_name in enumerate(df.columns):
            sanitized_columns[col_name] = unique_columns[i]

        # Escrever cabecalhos unicos
        for col_idx, col_name in enumerate(unique_columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Escrever dados usando índices numéricos para evitar erro de Series ambígua
        for row_idx in range(len(df)):
            for col_idx in range(len(df.columns)):
                value = df.iloc[row_idx, col_idx]
                value = self._convert_cell_value(value)
                ws.cell(row=row_idx + 2, column=col_idx + 1, value=value)

        # Última linha de dados
        last_data_row = len(df) + 1

        # Criar tabela SEM linha de totais
        tab_ref = f"A1:{get_column_letter(len(df.columns))}{last_data_row}"
        self._create_table(
            ws,
            table_name=settings.TABLE_ACESSOS,
            ref=tab_ref
        )

        # Linha de totais FORA da tabela (como células normais)
        total_row = last_data_row + 1
        nome_col_name = sanitized_columns.get('Nome', 'Nome')
        ws[f'A{total_row}'] = f"=SUBTOTAL(103,{settings.TABLE_ACESSOS}[{nome_col_name}])"
        ws[f'A{total_row}'].font = Font(bold=True)

        if 'Tempo_Minutos' in df.columns:
            tempo_col_idx = df.columns.get_loc('Tempo_Minutos') + 1
            tempo_col_letter = get_column_letter(tempo_col_idx)
            tempo_col_name = sanitized_columns.get('Tempo_Minutos', 'Tempo_Minutos')
            ws[f'{tempo_col_letter}{total_row}'] = (
                f"=SUBTOTAL(101,{settings.TABLE_ACESSOS}[{tempo_col_name}])/1440"
            )
            ws[f'{tempo_col_letter}{total_row}'].number_format = '[h]:mm'
            ws[f'{tempo_col_letter}{total_row}'].font = Font(bold=True)
        else:
            last_col_letter = get_column_letter(len(df.columns))
            retencao_col_name = sanitized_columns.get('Retenção (hh:mm)', 'Retencao_hh_mm')
            ws[f'{last_col_letter}{total_row}'] = (
                f"=SUBTOTAL(101,{settings.TABLE_ACESSOS}[{retencao_col_name}])"
            )
            ws[f'{last_col_letter}{total_row}'].font = Font(bold=True)

        # Ajustar larguras das colunas
        self._auto_adjust_column_widths(ws, max_width=50)

    def _create_inscritos_sheet(self, wb: Workbook, df: pd.DataFrame) -> None:
        """
        Cria planilha "Inscritos" com lista de inscritos.

        Args:
            wb: Workbook do openpyxl
            df: DataFrame com inscritos
        """
        ws = wb.create_sheet("Inscritos")

        # Sanitizar nomes das colunas
        sanitized_list = [self._sanitize_column_name(col) for col in df.columns]

        # Garantir unicidade dos nomes (evita corrupcao do XML)
        unique_columns = self._ensure_unique_column_names(sanitized_list)

        # Mapear nome original -> nome final unico
        sanitized_columns = {}
        for i, col_name in enumerate(df.columns):
            sanitized_columns[col_name] = unique_columns[i]

        # Escrever cabecalhos unicos
        for col_idx, col_name in enumerate(unique_columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Escrever dados usando índices numéricos para evitar erro de Series ambígua
        for row_idx in range(len(df)):
            for col_idx in range(len(df.columns)):
                value = df.iloc[row_idx, col_idx]
                value = self._convert_cell_value(value)
                ws.cell(row=row_idx + 2, column=col_idx + 1, value=value)

        # Última linha de dados
        last_data_row = len(df) + 1

        # Criar tabela SEM linha de totais
        tab_ref = f"A1:{get_column_letter(len(df.columns))}{last_data_row}"
        self._create_table(
            ws,
            table_name=settings.TABLE_INSCRITOS,
            ref=tab_ref
        )

        # Linha de totais FORA da tabela (como células normais)
        total_row = last_data_row + 1
        ws[f'A{total_row}'] = "Total"
        ws[f'A{total_row}'].font = Font(bold=True)
        last_col_letter = get_column_letter(len(df.columns))
        # Usar nome sanitizado
        data_cadastro_col = sanitized_columns.get('Data de Cadastro', 'Data_de_Cadastro')
        ws[f'{last_col_letter}{total_row}'] = (
            f"=SUBTOTAL(103,{settings.TABLE_INSCRITOS}[{data_cadastro_col}])"
        )
        ws[f'{last_col_letter}{total_row}'].font = Font(bold=True)

        # Ajustar larguras das colunas
        self._auto_adjust_column_widths(ws, max_width=50)

    def _create_enquete_sheet(
        self,
        wb: Workbook,
        df: pd.DataFrame,
        sheet_name: str,
        enquete_num: int
    ) -> None:
        """
        Cria planilha de enquete com tabela formatada.

        Colunas esperadas: Nome, Login, Pergunta, Resposta, Data

        Args:
            wb: Workbook do openpyxl
            df: DataFrame processado da enquete
            sheet_name: Nome da aba (ex: "Enquete 01")
            enquete_num: Número da enquete (para nome único da tabela)
        """
        ws = wb.create_sheet(sheet_name)

        # Sanitizar nomes de colunas
        raw_columns = df.columns.tolist()
        sanitized = [self._sanitize_column_name(c) for c in raw_columns]
        sanitized = self._ensure_unique_column_names(sanitized)
        sanitized_map = dict(zip(raw_columns, sanitized))

        # Cabeçalhos
        for col_idx, col_name in enumerate(raw_columns):
            ws.cell(row=1, column=col_idx + 1, value=sanitized_map[col_name])

        # Dados
        for row_idx, (_, row) in enumerate(df.iterrows()):
            for col_idx, col_name in enumerate(raw_columns):
                value = self._convert_cell_value(row[col_name])
                ws.cell(row=row_idx + 2, column=col_idx + 1, value=value)

        last_data_row = len(df) + 1

        # Nome único da tabela por número de enquete
        table_name = f"tblEnquete{enquete_num:02d}"
        tab_ref = f"A1:{get_column_letter(len(raw_columns))}{last_data_row}"
        self._create_table(ws, table_name=table_name, ref=tab_ref)

        # Total fora da tabela
        total_row = last_data_row + 1
        ws[f'A{total_row}'] = "Total"
        ws[f'A{total_row}'].font = Font(bold=True)
        # Contar pela coluna Nome (sempre primeira)
        ws[f'B{total_row}'] = f"=SUBTOTAL(103,{table_name}[{sanitized[1] if len(sanitized) > 1 else sanitized[0]}])"
        ws[f'B{total_row}'].font = Font(bold=True)

        self._auto_adjust_column_widths(ws, max_width=60)

    # Contador de IDs de tabela para garantir unicidade
    _table_id_counter = 0

    def _create_table(
        self,
        ws: Worksheet,
        table_name: str,
        ref: str
    ) -> None:
        """
        Cria tabela Excel com formatação padrão (SEM linha de totais).

        A linha de totais será adicionada como células normais fora da tabela
        para evitar problemas de corrupção do openpyxl com totalsRowShown.

        Args:
            ws: Worksheet onde criar a tabela
            table_name: Nome único da tabela
            ref: Referência da tabela (ex: "A1:C10")

        Example:
            >>> self._create_table(ws, "minhaTabela", "A1:D100")
        """
        # Incrementar contador de ID para garantir unicidade
        ExcelGenerator._table_id_counter += 1
        table_id = ExcelGenerator._table_id_counter

        # Criar tabela com ID explicito para evitar conflitos
        tab = Table(displayName=table_name, ref=ref, id=table_id)

        # Aplicar estilo
        style = TableStyleInfo(
            name=settings.EXCEL_TABLE_STYLE,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style

        # NÃO usar totalsRowShown - evita problemas de corrupção
        # Os totais são adicionados como células normais fora da tabela

        ws.add_table(tab)

    def _auto_adjust_column_widths(
        self,
        ws: Worksheet,
        max_width: int = 50
    ) -> None:
        """
        Ajusta automaticamente as larguras das colunas baseado no conteúdo.

        Helper reutilizável para ajustar larguras de todas as colunas.

        Args:
            ws: Worksheet para ajustar colunas
            max_width: Largura máxima permitida para uma coluna

        Example:
            >>> self._auto_adjust_column_widths(ws, max_width=60)
        """
        for column_cells in ws.columns:
            # Calcular largura baseada no maior conteúdo
            length = max(len(str(cell.value or '')) for cell in column_cells)
            # Adicionar margem e limitar ao máximo
            adjusted_width = min(length + 2, max_width)
            # Aplicar largura
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width
