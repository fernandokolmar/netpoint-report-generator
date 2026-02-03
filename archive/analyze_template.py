#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para analisar o template Excel de referência
"""

import openpyxl
from openpyxl.pivot.table import TableDefinition
import sys

def analyze_excel_template(file_path):
    """Analisar estrutura do template Excel"""
    print("=" * 80)
    print("ANÁLISE DO TEMPLATE EXCEL DE REFERÊNCIA")
    print("=" * 80)

    try:
        # Carregar workbook
        wb = openpyxl.load_workbook(file_path)

        print(f"\nPLANILHAS ENCONTRADAS: {len(wb.sheetnames)}")
        print("-" * 80)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\nPLANILHA: '{sheet_name}'")
            print(f"   Dimensões: {ws.dimensions}")

            # Verificar tabelas
            if hasattr(ws, '_tables') and ws._tables:
                print(f"   [OK] TABELAS ENCONTRADAS: {len(ws._tables)}")
                for table_name, table_ref in ws._tables.items():
                    print(f"      - Nome: {table_name}")
                    try:
                        # Tentar obter o objeto da tabela
                        if hasattr(table_ref, 'ref'):
                            print(f"        Referencia: {table_ref.ref}")
                            print(f"        Estilo: {table_ref.tableStyleInfo.name if table_ref.tableStyleInfo else 'N/A'}")
                        else:
                            print(f"        Referencia: {table_ref}")
                    except:
                        print(f"        Referencia: {table_ref}")
            else:
                print(f"   [X] Nenhuma tabela encontrada")

            # Verificar tabelas dinâmicas (pivot tables)
            if hasattr(ws, '_pivots') and ws._pivots:
                print(f"   [OK] TABELAS DINAMICAS (PIVOT TABLES): {len(ws._pivots)}")
                for pivot in ws._pivots:
                    print(f"      - Nome: {pivot.name if hasattr(pivot, 'name') else 'N/A'}")
                    print(f"        Cache ID: {pivot.cacheId if hasattr(pivot, 'cacheId') else 'N/A'}")
            else:
                print(f"   [X] Nenhuma tabela dinamica encontrada")

            # Verificar gráficos
            if hasattr(ws, '_charts') and ws._charts:
                print(f"   [OK] GRAFICOS ENCONTRADOS: {len(ws._charts)}")
                for chart in ws._charts:
                    print(f"      - Tipo: {type(chart).__name__}")
                    print(f"        Titulo: {chart.title if hasattr(chart, 'title') else 'N/A'}")
            else:
                print(f"   [X] Nenhum grafico encontrado")

            # Verificar formatação condicional
            if hasattr(ws, 'conditional_formatting') and ws.conditional_formatting:
                print(f"   [OK] FORMATACAO CONDICIONAL: {len(ws.conditional_formatting._cf_rules)}")

            # Mostrar algumas células com dados
            print(f"\n   PRIMEIRAS LINHAS:")
            max_row_display = min(5, ws.max_row)
            max_col_display = min(8, ws.max_column)

            for row_idx in range(1, max_row_display + 1):
                row_values = []
                for col_idx in range(1, max_col_display + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = str(cell.value)[:20] if cell.value else ''
                    row_values.append(value)
                print(f"      Linha {row_idx}: {' | '.join(row_values)}")

        print("\n" + "=" * 80)
        print("ANÁLISE CONCLUÍDA")
        print("=" * 80)

    except Exception as e:
        print(f"\n[X] ERRO: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    template_path = r"c:\Users\ferna\Desktop\App Estatisticas\Temple PRSA  - estatísitcas - 2025-11-18.xlsx"
    analyze_excel_template(template_path)
