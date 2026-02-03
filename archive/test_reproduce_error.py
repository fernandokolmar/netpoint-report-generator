"""Reproduce the exact error happening in the Excel generator"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Simulate the actual data that might be causing issues
def test_series_ambiguity():
    """Test for 'truth value of Series is ambiguous' error"""

    # Create sample DataFrame similar to what data_processor creates
    df = pd.DataFrame({
        'Nome': ['Alice', 'Bob', 'Charlie'],
        'Celular': ['123', '456', '789'],
        'Tempo_Minutos': [90.5, 120.0, 45.3]
    })

    print("DataFrame info:")
    print(df.dtypes)
    print("\nDataFrame:")
    print(df)

    wb = Workbook()
    ws = wb.active

    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Try the approach used in excel_generator.py
    print("\n\nTest 1: Using df.iloc[row_idx][col_name]")
    try:
        for row_idx in range(len(df)):
            for col_idx, col_name in enumerate(df.columns, 1):
                value = df.iloc[row_idx][col_name]
                print(f"Row {row_idx}, Col {col_name}: type={type(value)}, value={value}")

                # This is where the error might occur
                if pd.isna(value):
                    print("  -> Setting to None (was NaN)")
                    value = None
                elif hasattr(value, 'item'):
                    print(f"  -> Has .item() method, converting from {type(value)}")
                    value = value.item()

                ws.cell(row=row_idx + 2, column=col_idx, value=value)
        print("SUCCESS: No error with approach 1")
    except Exception as e:
        print(f"ERROR in approach 1: {e}")
        import traceback
        traceback.print_exc()

    # Test approach 2: Direct column access
    print("\n\nTest 2: Using df.iloc[row_idx, col_idx]")
    try:
        for row_idx in range(len(df)):
            for col_idx in range(len(df.columns)):
                value = df.iloc[row_idx, col_idx]
                col_name = df.columns[col_idx]
                print(f"Row {row_idx}, Col {col_name}: type={type(value)}, value={value}")

                if pd.isna(value):
                    print("  -> Setting to None (was NaN)")
                    value = None
                elif hasattr(value, 'item'):
                    print(f"  -> Has .item() method, converting from {type(value)}")
                    value = value.item()

                ws.cell(row=row_idx + 2, column=col_idx + 1, value=value)
        print("SUCCESS: No error with approach 2")
    except Exception as e:
        print(f"ERROR in approach 2: {e}")
        import traceback
        traceback.print_exc()

def test_with_problematic_data():
    """Test with data that might cause issues"""

    print("\n\n" + "="*60)
    print("TEST WITH PROBLEMATIC DATA")
    print("="*60)

    # Create DataFrame with various problematic types
    df = pd.DataFrame({
        'Nome': ['Alice', 'Bob', np.nan],  # Contains NaN
        'Value': [1, 2, 3],  # Regular integers
        'Float': [1.5, 2.5, 3.5],  # Floats
        'Mixed': ['text', 123, None],  # Mixed types
    })

    print("DataFrame:")
    print(df)
    print("\nDtypes:")
    print(df.dtypes)

    wb = Workbook()
    ws = wb.active

    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Use the excel_generator approach
    print("\nWriting data:")
    for row_idx in range(len(df)):
        print(f"\n  Row {row_idx}:")
        for col_idx, col_name in enumerate(df.columns, 1):
            value = df.iloc[row_idx][col_name]
            print(f"    {col_name}: {repr(value)} (type: {type(value).__name__})")

            # Check if this causes the ambiguous truth value error
            try:
                if pd.isna(value):
                    print(f"      -> is NaN, setting to None")
                    value = None
                elif hasattr(value, 'item'):
                    print(f"      -> has .item(), converting")
                    value = value.item()

                ws.cell(row=row_idx + 2, column=col_idx, value=value)
            except ValueError as e:
                if "truth value" in str(e).lower():
                    print(f"      -> ERROR: Truth value ambiguity!")
                    print(f"         This is the bug! Value type: {type(value)}")
                    raise
                else:
                    raise

def test_series_vs_scalar():
    """Test to understand when iloc returns Series vs scalar"""

    print("\n\n" + "="*60)
    print("TEST: Series vs Scalar")
    print("="*60)

    df = pd.DataFrame({
        'A': [1, 2, 3],
        'B': [4, 5, 6]
    })

    print("Test 1: df.iloc[0] (single row index)")
    result1 = df.iloc[0]
    print(f"  Type: {type(result1)}")
    print(f"  Value: {result1}")
    print(f"  This is a Series!")

    print("\nTest 2: df.iloc[0]['A'] (chained indexing)")
    result2 = df.iloc[0]['A']
    print(f"  Type: {type(result2)}")
    print(f"  Value: {result2}")
    print(f"  This is a scalar!")

    print("\nTest 3: df.iloc[0, 0] (tuple indexing)")
    result3 = df.iloc[0, 0]
    print(f"  Type: {type(result3)}")
    print(f"  Value: {result3}")
    print(f"  This is a scalar!")

    print("\nConclusion: df.iloc[row_idx][col_name] should return scalar, not Series")
    print("Unless... there's some weird edge case?")

if __name__ == "__main__":
    test_series_ambiguity()
    test_with_problematic_data()
    test_series_vs_scalar()
