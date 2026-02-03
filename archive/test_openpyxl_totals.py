from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import zipfile

# Create test workbook
wb = Workbook()
ws = wb.active

# Write data
ws['A1'] = 'Name'
ws['B1'] = 'Value'
ws['A2'] = 'Alice'
ws['B2'] = 1
ws['A3'] = 'Bob'
ws['B3'] = 2

# Create table WITH totals row
tab = Table(displayName="TestTable", ref="A1:B4")
tab.totalsRowShown = True
style = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)
tab.tableStyleInfo = style
ws.add_table(tab)

# Save
wb.save("test_totals_shown.xlsx")

# Read the XML
z = zipfile.ZipFile("test_totals_shown.xlsx")
xml = z.read('xl/tables/table1.xml').decode('utf-8')
print(xml)
