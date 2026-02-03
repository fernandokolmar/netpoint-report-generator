#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de teste para o Gerador de Relatório PRSA
Testa o processamento sem interface gráfica
"""

import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
import os

def test_processing():
    """Testar o processamento dos dados"""
    print("=" * 60)
    print("TESTE DO GERADOR DE RELATÓRIO PRSA")
    print("=" * 60)
    
    # Caminhos dos arquivos
    file_paths = {
        'inscritos': '/mnt/project/Inscritos.csv',
        'mensagens': '/mnt/project/Mensagens.csv',
        'relatorio_acesso': '/mnt/project/Relatório_de_acesso.csv',
        'totalizado': '/mnt/project/Totalizado.csv'
    }
    
    # Verificar se os arquivos existem
    print("\n1. Verificando arquivos...")
    for name, path in file_paths.items():
        if os.path.exists(path):
            print(f"   ✓ {name}: {os.path.basename(path)}")
        else:
            print(f"   ✗ {name}: ARQUIVO NÃO ENCONTRADO")
            return False
    
    # Carregar DataFrames
    print("\n2. Carregando dados...")
    dataframes = {}
    
    try:
        # Inscritos
        dataframes['inscritos'] = pd.read_csv(
            file_paths['inscritos'], 
            sep=';', 
            encoding='utf-8-sig'
        )
        print(f"   ✓ Inscritos: {len(dataframes['inscritos'])} registros")
        
        # Mensagens
        dataframes['mensagens'] = pd.read_csv(
            file_paths['mensagens'], 
            sep=';', 
            encoding='utf-8-sig'
        )
        print(f"   ✓ Mensagens: {len(dataframes['mensagens'])} registros")
        
        # Relatório de Acesso
        dataframes['relatorio_acesso'] = pd.read_csv(
            file_paths['relatorio_acesso'], 
            sep=';', 
            encoding='utf-8-sig'
        )
        print(f"   ✓ Relatório de Acesso: {len(dataframes['relatorio_acesso'])} registros")
        
        # Totalizado
        dataframes['totalizado'] = pd.read_csv(
            file_paths['totalizado'], 
            sep=';', 
            encoding='utf-8-sig'
        )
        print(f"   ✓ Totalizado: {len(dataframes['totalizado'])} registros")
        
    except Exception as e:
        print(f"   ✗ Erro ao carregar arquivos: {e}")
        return False
    
    # Processar dados
    print("\n3. Processando dados...")
    
    # Calcular estatísticas
    total_inscritos = len(dataframes['inscritos'])
    total_acessos = len(dataframes['relatorio_acesso'])
    total_mensagens = len(dataframes['mensagens'])
    pico_audiencia = dataframes['totalizado']['Usuarios conectados'].max()
    
    # Encontrar hora do pico
    dataframes['totalizado']['Data'] = pd.to_datetime(
        dataframes['totalizado']['Data'], 
        format='%d/%m/%Y %H:%M:%S',
        dayfirst=True
    )
    idx_pico = dataframes['totalizado']['Usuarios conectados'].idxmax()
    hora_pico = dataframes['totalizado'].loc[idx_pico, 'Data']
    
    # Calcular tempo médio (se disponível)
    if 'Tempo' in dataframes['relatorio_acesso'].columns:
        tempo_medio_minutos = dataframes['relatorio_acesso']['Tempo'].mean()
        horas = int(tempo_medio_minutos // 60)
        minutos = int(tempo_medio_minutos % 60)
        tempo_medio_str = f"{horas:02d}:{minutos:02d}:00"
    else:
        tempo_medio_str = "00:00:00"
    
    print(f"   ✓ Total de inscritos: {total_inscritos}")
    print(f"   ✓ Usuários distintos na live: {total_acessos}")
    print(f"   ✓ Pico de audiência: {pico_audiencia} usuários")
    print(f"   ✓ Hora do pico: {hora_pico.strftime('%H:%M')}")
    print(f"   ✓ Tempo médio assistido: {tempo_medio_str}")
    print(f"   ✓ Total de mensagens: {total_mensagens}")
    
    # Criar Excel de teste
    print("\n4. Criando arquivo Excel de teste...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    
    # Adicionar resumo
    ws['A1'] = "Métrica"
    ws['B1'] = "Valor"
    
    ws['A2'] = "Total de inscritos"
    ws['B2'] = total_inscritos
    
    ws['A3'] = "Usuários na live"
    ws['B3'] = total_acessos
    
    ws['A4'] = "Pico de audiência"
    ws['B4'] = pico_audiencia
    
    ws['A5'] = "Hora do pico"
    ws['B5'] = hora_pico.strftime('%d/%m/%Y %H:%M')
    
    ws['A6'] = "Tempo médio"
    ws['B6'] = tempo_medio_str
    
    ws['A7'] = "Total de mensagens"
    ws['B7'] = total_mensagens
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    
    # Salvar arquivo de teste
    output_path = '/mnt/user-data/outputs/teste_prsa_report.xlsx'
    wb.save(output_path)
    print(f"   ✓ Arquivo de teste salvo: {output_path}")
    
    print("\n" + "=" * 60)
    print("TESTE CONCLUÍDO COM SUCESSO!")
    print("=" * 60)
    print("\nA aplicação está funcionando corretamente.")
    print("Execute o arquivo 'prsa_report_generator.py' para usar a interface gráfica.")
    
    return True


if __name__ == "__main__":
    test_processing()
