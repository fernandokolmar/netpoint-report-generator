"""
Script de teste para verificar se as fórmulas Excel estão corretas.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Criar workbook de teste
wb = Workbook()
ws = wb.active
ws.title = "Teste"

# Dados de teste
data = {
    'Nome': ['João', 'Maria', 'Pedro'],
    'Retenção (hh:mm)': ['1:30:00', '2:15:00', '0:45:00'],
    'Tempo_Minutos': [90.0, 135.0, 45.0]
}

df = pd.DataFrame(data)

# Adicionar headers
for c_idx, col_name in enumerate(df.columns, 1):
    ws.cell(row=1, column=c_idx, value=col_name)

# Adicionar dados
for r_idx, row in df.iterrows():
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx + 2, column=c_idx, value=value)

# Criar tabela Excel (com linha totalizadora)
tab_ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 2}"
print(f"Referência da tabela: {tab_ref}")

tab = Table(displayName='Relatorio_de_acesso12', ref=tab_ref)
tab.totalsRowShown = True

style = TableStyleInfo(
    name="TableStyleMedium12",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)
tab.tableStyleInfo = style

ws.add_table(tab)

# Adicionar fórmulas na linha totalizadora
total_row = len(df) + 2

# Fórmula SUBTOTAL na coluna Nome
ws[f'A{total_row}'] = "=SUBTOTAL(103,Relatorio_de_acesso12[Nome])"

# Fórmula SUBTOTAL na coluna Tempo_Minutos
tempo_col_idx = df.columns.get_loc('Tempo_Minutos') + 1
tempo_col_letter = get_column_letter(tempo_col_idx)
print(f"Coluna Tempo_Minutos: {tempo_col_letter}")

ws[f'{tempo_col_letter}{total_row}'] = "=SUBTOTAL(101,Relatorio_de_acesso12[Tempo_Minutos])"
ws[f'{tempo_col_letter}{total_row}'].number_format = '[h]:mm'

# Criar célula para teste AVERAGE (fora da tabela)
ws['E1'] = 'Tempo médio (AVERAGE):'
ws['F1'] = "=AVERAGE(Relatorio_de_acesso12[Tempo_Minutos])"
ws['F1'].number_format = '[h]:mm'

# Salvar arquivo
output_path = "c:/Users/ferna/Desktop/teste_formulas.xlsx"
wb.save(output_path)

print(f"Arquivo de teste salvo: {output_path}")
print("\nVerifique se:")
print("1. O arquivo abre sem erro")
print("2. A fórmula AVERAGE na célula F1 funciona")
print("3. A fórmula SUBTOTAL na linha totalizadora funciona")
print("4. O formato [h]:mm exibe corretamente")
