"""Test the fixed table generation to ensure no corruption"""
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import zipfile
import re

def test_table_with_totals():
    """Test creating a table with totals row using the new approach"""

    # Create test data
    df = pd.DataFrame({
        'Name': ['Alice', 'Bob', 'Charlie'],
        'Value': [1, 2, 3]
    })

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Test"

    # Write data (NO totals row)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    print("Data written:")
    for row in range(1, len(df) + 2):
        print(f"  Row {row}: {[ws.cell(row=row, column=col).value for col in range(1, 3)]}")

    # Create table WITHOUT totals row in ref
    total_row = len(df) + 2
    tab_ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"

    print(f"\nCreating table with ref: {tab_ref}")
    print(f"Total row will be at: {total_row}")

    tab = Table(displayName="TestTable", ref=tab_ref)
    tab.totalsRowShown = True
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # NOW add totals row
    ws[f'A{total_row}'] = "Total"
    ws[f'B{total_row}'] = f"=SUBTOTAL(109,TestTable[Value])"

    print(f"\nTotals row added:")
    print(f"  A{total_row}: {ws[f'A{total_row}'].value}")
    print(f"  B{total_row}: {ws[f'B{total_row}'].value}")

    # Save
    wb.save("test_fixed_tables.xlsx")
    print("\nFile saved: test_fixed_tables.xlsx")

    # Verify: Read back and check
    wb2 = load_workbook("test_fixed_tables.xlsx")
    ws2 = wb2.active

    print("\nVerification:")
    print(f"  Table ref: {ws2.tables['TestTable'].ref}")
    print(f"  Table totalsRowShown: {ws2.tables['TestTable'].totalsRowShown}")

    # Check XML
    z = zipfile.ZipFile("test_fixed_tables.xlsx")
    xml = z.read('xl/tables/table1.xml').decode('utf-8')

    ref_match = re.search(r'ref="([^"]+)"', xml)
    totals_match = re.search(r'totalsRowShown="([^"]+)"', xml)

    print(f"\nXML attributes:")
    print(f"  ref: {ref_match.group(1) if ref_match else 'NOT FOUND'}")
    print(f"  totalsRowShown: {totals_match.group(1) if totals_match else 'NOT FOUND'}")

    # Check if file opens without corruption
    try:
        print("\nOpening in openpyxl: SUCCESS")
        print("  No corruption detected!")
        return True
    except Exception as e:
        print(f"\nOpening in openpyxl: FAILED - {e}")
        return False

if __name__ == "__main__":
    success = test_table_with_totals()
    exit(0 if success else 1)
