#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Análise COMPLETA E PROFUNDA do template
"""

import openpyxl
from openpyxl.utils import get_column_letter

def deep_analysis():
    template_path = r"c:\Users\ferna\Desktop\App Estatisticas\Temple PRSA  - estatísitcas - 2025-11-18.xlsx"

    print("=" * 100)
    print("ANALISE COMPLETA DO TEMPLATE")
    print("=" * 100)

    wb = openpyxl.load_workbook(template_path, data_only=False)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*100}")
        print(f"PLANILHA: {sheet_name}")
        print(f"{'='*100}")
        print(f"Dimensoes: {ws.dimensions}")
        print(f"Total de linhas: {ws.max_row}")
        print(f"Total de colunas: {ws.max_column}")

        # Mostrar cabeçalhos (linha 1)
        print(f"\nCABECALHOS (Linha 1):")
        for col_idx in range(1, min(20, ws.max_column + 1)):
            cell = ws.cell(row=1, column=col_idx)
            col_letter = get_column_letter(col_idx)
            if cell.value:
                print(f"  {col_letter}1: {str(cell.value)[:40]}")

        # Mostrar linha 2 (primeira linha de dados)
        print(f"\nPRIMEIRA LINHA DE DADOS (Linha 2):")
        for col_idx in range(1, min(20, ws.max_column + 1)):
            cell = ws.cell(row=2, column=col_idx)
            col_letter = get_column_letter(col_idx)
            if cell.value:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    print(f"  {col_letter}2: [FORMULA] {str(cell.value)[:60]}")
                else:
                    print(f"  {col_letter}2: {str(cell.value)[:40]}")

        # Mostrar ÚLTIMA linha (totalizadora)
        print(f"\nLINHA TOTALIZADORA (Linha {ws.max_row}):")
        for col_idx in range(1, min(20, ws.max_column + 1)):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            col_letter = get_column_letter(col_idx)
            if cell.value:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    print(f"  {col_letter}{ws.max_row}: [FORMULA] {str(cell.value)[:60]}")
                else:
                    print(f"  {col_letter}{ws.max_row}: {str(cell.value)[:40]}")

        # Informações sobre tabelas
        if hasattr(ws, '_tables') and ws._tables:
            print(f"\nTABELAS:")
            for table_name in ws._tables:
                print(f"  - {table_name}")

        # Informações sobre gráficos
        if hasattr(ws, '_charts') and ws._charts:
            print(f"\nGRAFICOS: {len(ws._charts)}")
            for idx, chart in enumerate(ws._charts):
                print(f"  Grafico {idx+1}:")
                print(f"    Tipo: {type(chart).__name__}")
                print(f"    Ancora: {chart.anchor}")
                if hasattr(chart, 'width'):
                    print(f"    Largura: {chart.width}")
                if hasattr(chart, 'height'):
                    print(f"    Altura: {chart.height}")

if __name__ == "__main__":
    try:
        deep_analysis()
    except Exception as e:
        print(f"Erro: {e}")
        import traceback
        traceback.print_exc()
