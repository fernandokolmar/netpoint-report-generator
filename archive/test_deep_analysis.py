"""
DEEP ANALYSIS: Find the ROOT CAUSE of table corruption
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import zipfile
import re

def test_hypothesis_empty_cells():
    """
    HYPOTHESIS 1: The corruption happens when cells in the totals row
    are empty BEFORE the table is created, then filled AFTER
    """
    print("="*70)
    print("HYPOTHESIS 1: Empty cells in totals row before table creation")
    print("="*70)

    wb = Workbook()
    ws = wb.active
    ws.title = "Test1"

    # Write data
    ws['A1'] = 'Name'
    ws['B1'] = 'Value'
    ws['A2'] = 'Alice'
    ws['B2'] = 1
    ws['A3'] = 'Bob'
    ws['B3'] = 2

    # Check row 4 (totals row) - should be empty
    print(f"\nRow 4 before table: A4={repr(ws['A4'].value)}, B4={repr(ws['B4'].value)}")

    # Create table with row 4 as totals
    tab = Table(displayName="Test1", ref="A1:B4")
    tab.totalsRowShown = True
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium12",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    ws.add_table(tab)

    print(f"Table created with ref=A1:B4, totalsRowShown=True")

    # NOW write to totals row
    ws['A4'] = "Total"
    ws['B4'] = "=SUBTOTAL(109,Test1[Value])"
    print(f"Wrote to row 4: A4={repr(ws['A4'].value)}, B4={repr(ws['B4'].value)}")

    # Save
    wb.save("test_hypothesis1.xlsx")
    print("\nSaved as: test_hypothesis1.xlsx")

    return check_for_corruption("test_hypothesis1.xlsx")

def test_hypothesis_column_headers():
    """
    HYPOTHESIS 2: The corruption happens due to how column headers are written
    """
    print("\n" + "="*70)
    print("HYPOTHESIS 2: Column header writing method")
    print("="*70)

    wb = Workbook()

    # Method 1: Direct cell assignment (like Retencao - WORKS)
    ws1 = wb.create_sheet("Method1_Direct")
    ws1['A1'] = 'Name'
    ws1['B1'] = 'Value'
    ws1['A2'] = 'Alice'
    ws1['B2'] = 1

    tab1 = Table(displayName="DirectHeaders", ref="A1:B3")
    tab1.totalsRowShown = True
    tab1.tableStyleInfo = TableStyleInfo(name="TableStyleMedium12", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws1.add_table(tab1)
    ws1['A3'] = "Total"
    ws1['B3'] = "=SUBTOTAL(109,DirectHeaders[Value])"

    # Method 2: ws.cell() with iteration (like Mensagens - BREAKS)
    ws2 = wb.create_sheet("Method2_Cell")
    df = pd.DataFrame({'Name': ['Alice'], 'Value': [1]})

    for col_idx, col_name in enumerate(df.columns, 1):
        ws2.cell(row=1, column=col_idx, value=col_name)

    for row_idx in range(len(df)):
        for col_idx, col_name in enumerate(df.columns, 1):
            value = df.iloc[row_idx][col_name]
            ws2.cell(row=row_idx + 2, column=col_idx, value=value)

    tab2 = Table(displayName="CellHeaders", ref="A1:B3")
    tab2.totalsRowShown = True
    tab2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium12", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws2.add_table(tab2)
    ws2['A3'] = "Total"
    ws2['B3'] = "=SUBTOTAL(109,CellHeaders[Value])"

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    wb.save("test_hypothesis2.xlsx")
    print("\nSaved as: test_hypothesis2.xlsx")

    return check_for_corruption("test_hypothesis2.xlsx")

def test_hypothesis_table_names():
    """
    HYPOTHESIS 3: Special characters or spaces in table names cause corruption
    """
    print("\n" + "="*70)
    print("HYPOTHESIS 3: Table naming issues")
    print("="*70)

    wb = Workbook()

    test_names = [
        ("Simple", "simple_table"),
        ("Underscore", "table_with_underscore"),
        ("Number", "Relatorio_de_acesso12"),  # Like TABLE_ACESSOS
        ("AllCaps", "TABLE_MENSAGENS"),
        ("LowerCase", "mensagens"),
    ]

    for sheet_name, table_name in test_names:
        ws = wb.create_sheet(sheet_name)
        ws['A1'] = 'Name'
        ws['B1'] = 'Value'
        ws['A2'] = 'Alice'
        ws['B2'] = 1

        tab = Table(displayName=table_name, ref="A1:B3")
        tab.totalsRowShown = True
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium12", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws.add_table(tab)

        ws['A3'] = "Total"
        ws['B3'] = f"=SUBTOTAL(109,{table_name}[Value])"

        print(f"  Created table: {table_name}")

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    wb.save("test_hypothesis3.xlsx")
    print("\nSaved as: test_hypothesis3.xlsx")

    return check_for_corruption("test_hypothesis3.xlsx")

def test_hypothesis_formula_references():
    """
    HYPOTHESIS 4: Formula references to the table cause corruption
    """
    print("\n" + "="*70)
    print("HYPOTHESIS 4: Formula reference issues")
    print("="*70)

    wb = Workbook()

    # Test 1: No formulas in totals row
    ws1 = wb.create_sheet("NoFormula")
    ws1['A1'] = 'Name'
    ws1['B1'] = 'Value'
    ws1['A2'] = 'Alice'
    ws1['B2'] = 1

    tab1 = Table(displayName="NoFormulaTable", ref="A1:B3")
    tab1.totalsRowShown = True
    tab1.tableStyleInfo = TableStyleInfo(name="TableStyleMedium12", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws1.add_table(tab1)

    ws1['A3'] = "Total"
    ws1['B3'] = "Static Text"  # NO FORMULA

    # Test 2: With SUBTOTAL formula
    ws2 = wb.create_sheet("WithFormula")
    ws2['A1'] = 'Name'
    ws2['B1'] = 'Value'
    ws2['A2'] = 'Alice'
    ws2['B2'] = 1

    tab2 = Table(displayName="WithFormulaTable", ref="A1:B3")
    tab2.totalsRowShown = True
    tab2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium12", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws2.add_table(tab2)

    ws2['A3'] = "Total"
    ws2['B3'] = "=SUBTOTAL(109,WithFormulaTable[Value])"

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    wb.save("test_hypothesis4.xlsx")
    print("\nSaved as: test_hypothesis4.xlsx")

    return check_for_corruption("test_hypothesis4.xlsx")

def check_for_corruption(filename):
    """Check if Excel file has corruption markers in XML"""
    print(f"\n  Checking {filename} for corruption...")

    with zipfile.ZipFile(filename, 'r') as z:
        # Check for recoveredErrors.xml (appears after Excel repair)
        if 'xl/recoveredErrors.xml' in z.namelist():
            print("    ❌ CORRUPTION FOUND: recoveredErrors.xml exists")
            return False

        # Check table XML for issues
        table_files = [f for f in z.namelist() if f.startswith('xl/tables/table') and f.endswith('.xml')]

        for table_file in table_files:
            xml = z.read(table_file).decode('utf-8')

            # Check for missing or incorrect attributes
            if not re.search(r'totalsRowCount="1"', xml):
                print(f"    ⚠️  WARNING: {table_file} missing totalsRowCount=\"1\"")

            if not re.search(r'totalsRowShown="1"', xml):
                print(f"    ❌ ERROR: {table_file} missing totalsRowShown=\"1\"")
                return False

            # Extract ref to check range
            ref_match = re.search(r'ref="([^"]+)"', xml)
            if ref_match:
                ref = ref_match.group(1)
                # Parse ref to check if last row matches autoFilter
                print(f"    ✓ {table_file}: ref={ref}")

    print(f"    ✓ No obvious corruption markers found")
    return True

def main():
    print("\n" + "="*70)
    print(" COMPREHENSIVE ROOT CAUSE ANALYSIS")
    print("="*70)

    results = {}

    results['hypothesis1'] = test_hypothesis_empty_cells()
    results['hypothesis2'] = test_hypothesis_column_headers()
    results['hypothesis3'] = test_hypothesis_table_names()
    results['hypothesis4'] = test_hypothesis_formula_references()

    print("\n" + "="*70)
    print(" RESULTS SUMMARY")
    print("="*70)

    for hypothesis, passed in results.items():
        status = "✓ PASSED" if passed else "❌ FAILED"
        print(f"  {hypothesis}: {status}")

    print("\n" + "="*70)
    print(" NEXT STEPS:")
    print("="*70)
    print("  1. Open each test file in Excel")
    print("  2. Check which ones show corruption error")
    print("  3. Compare against openpyxl's assessment")
    print("  4. Identify the root cause")
    print("="*70)

if __name__ == "__main__":
    main()
