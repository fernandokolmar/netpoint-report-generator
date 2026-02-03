"""
Test exact table creation pattern used in excel_generator.py
to identify what causes corruption in table2, table4, table5
"""
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import zipfile
import xml.etree.ElementTree as ET

def create_working_table(wb, sheet_name, table_name):
    """Create a table using the WORKING pattern from _create_retencao_sheet"""
    print(f"\n{'='*60}")
    print(f"Creating WORKING table: {table_name} in sheet {sheet_name}")
    print(f"{'='*60}")

    ws = wb.create_sheet(sheet_name)

    # Sample data
    df = pd.DataFrame({
        'Data': ['10:00', '10:01', '10:02'],
        'Usuarios conectados': [100, 150, 120]
    })

    # Write headers
    ws['A1'] = 'Data'
    ws['B1'] = 'Usuários conectados'
    ws['C1'] = 'Max'

    # Write data using iterrows (like Retencao)
    for i, (idx, row) in enumerate(df.iterrows()):
        row_num = i + 2
        ws[f'A{row_num}'] = row['Data']
        ws[f'B{row_num}'] = row['Usuarios conectados']

    # Create table INCLUDING totals row in ref
    tab_ref = f"A1:C{len(df) + 2}"
    print(f"Table ref: {tab_ref}")
    print(f"Last data row: {len(df) + 1}")
    print(f"Totals row: {len(df) + 2}")

    tab = Table(displayName=table_name, ref=tab_ref)
    tab.totalsRowShown = True
    style = TableStyleInfo(
        name="TableStyleMedium12",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Add formula to column C
    for i in range(len(df)):
        row_num = i + 2
        ws[f'C{row_num}'] = f'=IF({table_name}[[#This Row],[Usuários conectados]]=MAX({table_name}[Usuários conectados]),{table_name}[[#This Row],[Usuários conectados]],NA())'

    return ws

def create_broken_table(wb, sheet_name, table_name):
    """Create a table using the BROKEN pattern from _create_mensagens_sheet"""
    print(f"\n{'='*60}")
    print(f"Creating BROKEN table: {table_name} in sheet {sheet_name}")
    print(f"{'='*60}")

    ws = wb.create_sheet(sheet_name)

    # Sample data
    df = pd.DataFrame({
        'Data': ['10:00', '10:01', '10:02'],
        'Nome': ['Alice', 'Bob', 'Charlie'],
        'Mensagem': ['Hello', 'World', 'Test']
    })

    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Write data using df.iloc (like Mensagens/Acessos/Inscritos)
    for row_idx in range(len(df)):
        for col_idx, col_name in enumerate(df.columns, 1):
            value = df.iloc[row_idx][col_name]
            if pd.isna(value):
                value = None
            elif hasattr(value, 'item'):
                value = value.item()
            ws.cell(row=row_idx + 2, column=col_idx, value=value)

    # Totals row position
    total_row = len(df) + 2

    # Create table
    tab_ref = f"A1:{get_column_letter(len(df.columns))}{total_row}"
    print(f"Table ref: {tab_ref}")
    print(f"Last data row: {len(df) + 1}")
    print(f"Totals row: {total_row}")

    tab = Table(displayName=table_name, ref=tab_ref)
    tab.totalsRowShown = True
    style = TableStyleInfo(
        name="TableStyleMedium12",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Add formulas to totals row
    ws[f'A{total_row}'] = "Total"
    last_col = len(df.columns)
    ws.cell(row=total_row, column=last_col).value = f"=SUBTOTAL(103,{table_name}[Mensagem])"

    return ws

def analyze_xml_structure(filename):
    """Analyze the XML structure of tables to find differences"""
    print(f"\n{'='*60}")
    print(f"Analyzing XML structure of {filename}")
    print(f"{'='*60}")

    with zipfile.ZipFile(filename, 'r') as z:
        # List all table files
        table_files = [f for f in z.namelist() if f.startswith('xl/tables/table') and f.endswith('.xml')]
        print(f"\nFound {len(table_files)} tables:")

        for table_file in table_files:
            print(f"\n  {table_file}:")
            xml_content = z.read(table_file).decode('utf-8')

            # Parse XML
            root = ET.fromstring(xml_content)

            # Get namespace
            ns = {'main': root.tag.split('}')[0].strip('{')}

            # Extract key attributes
            ref = root.get('ref')
            display_name = root.get('displayName')
            totals_shown = root.get('totalsRowShown')
            totals_count = root.get('totalsRowCount')

            print(f"    displayName: {display_name}")
            print(f"    ref: {ref}")
            print(f"    totalsRowShown: {totals_shown}")
            print(f"    totalsRowCount: {totals_count}")

            # Check autoFilter
            auto_filter = root.find('.//main:autoFilter', ns)
            if auto_filter is not None:
                filter_ref = auto_filter.get('ref')
                print(f"    autoFilter ref: {filter_ref}")
            else:
                print(f"    autoFilter: None")

            # Check table columns
            table_columns = root.find('.//main:tableColumns', ns)
            if table_columns is not None:
                count = table_columns.get('count')
                print(f"    tableColumns count: {count}")

                # Check each column for totalsRowFunction
                for col in table_columns.findall('.//main:tableColumn', ns):
                    col_name = col.get('name')
                    totals_func = col.find('.//main:totalsRowFunction', ns)
                    totals_formula = col.get('totalsRowFormula')

                    if totals_func is not None or totals_formula is not None:
                        print(f"      Column '{col_name}':")
                        if totals_func is not None:
                            print(f"        totalsRowFunction: {totals_func.get('function')}")
                        if totals_formula:
                            print(f"        totalsRowFormula: {totals_formula}")

def compare_working_vs_broken():
    """Create both types and compare their XML"""

    # Create workbook with both types
    wb = Workbook()
    wb.remove(wb.active)

    # Create working table (like Retencao)
    create_working_table(wb, "Working", "TABLE_WORKING")

    # Create broken table (like Mensagens)
    create_broken_table(wb, "Broken", "TABLE_BROKEN")

    # Save
    filename = "test_compare_tables.xlsx"
    wb.save(filename)
    print(f"\n\nFile saved: {filename}")

    # Analyze
    analyze_xml_structure(filename)

    # Try to open in Excel (simulation - check for corruption)
    print(f"\n{'='*60}")
    print("Attempting to reload file...")
    print(f"{'='*60}")

    try:
        wb2 = load_workbook(filename)
        print("SUCCESS: File loaded without errors by openpyxl")

        # Check if tables exist
        for sheet_name in wb2.sheetnames:
            ws = wb2[sheet_name]
            print(f"\nSheet '{sheet_name}':")
            print(f"  Tables: {list(ws.tables.keys())}")
            for table_name, table in ws.tables.items():
                print(f"    {table_name}: ref={table.ref}, totalsRowShown={table.totalsRowShown}")

    except Exception as e:
        print(f"ERROR loading file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    compare_working_vs_broken()
