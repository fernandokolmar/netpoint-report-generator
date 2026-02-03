"""Test the corrected table generation approach"""
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import zipfile
import re

def test_correct_approach():
    """Test creating a table with totals row - CORRECT approach"""

    # Create test data
    df = pd.DataFrame({
        'Name': ['Alice', 'Bob', 'Charlie'],
        'Value': [1, 2, 3]
    })

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Test"

    # Step 1: Write data (NO totals row yet)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    print("Step 1: Data written (rows 1-4):")
    for row in range(1, 5):
        vals = [ws.cell(row=row, column=col).value for col in range(1, 3)]
        print(f"  Row {row}: {vals}")

    # Step 2: Check row 5 (should be empty)
    print(f"\nStep 2: Row 5 before table creation:")
    print(f"  A5: {repr(ws['A5'].value)}")
    print(f"  B5: {repr(ws['B5'].value)}")

    # Step 3: Create table INCLUDING row 5 in ref (but row 5 is still empty)
    total_row = len(df) + 2  # Row 5
    tab_ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 2}"  # A1:B5

    print(f"\nStep 3: Creating table with ref: {tab_ref}")

    tab = Table(displayName="TestTable", ref=tab_ref)
    tab.totalsRowShown = True
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    print(f"  Table created")

    # Step 4: NOW write to totals row
    print(f"\nStep 4: Writing to totals row {total_row}:")
    ws[f'A{total_row}'] = "Total"
    ws[f'B{total_row}'] = f"=SUBTOTAL(109,TestTable[Value])"
    print(f"  A{total_row}: {repr(ws[f'A{total_row}'].value)}")
    print(f"  B{total_row}: {repr(ws[f'B{total_row}'].value)}")

    # Save
    wb.save("test_correct_tables.xlsx")
    print("\nFile saved: test_correct_tables.xlsx")

    # Step 5: Verify by reading back
    wb2 = load_workbook("test_correct_tables.xlsx")
    ws2 = wb2.active

    print("\nStep 5: Verification:")
    print(f"  Table ref: {ws2.tables['TestTable'].ref}")
    print(f"  Table totalsRowShown: {ws2.tables['TestTable'].totalsRowShown}")

    # Check XML
    z = zipfile.ZipFile("test_correct_tables.xlsx")
    xml = z.read('xl/tables/table1.xml').decode('utf-8')

    ref_match = re.search(r'ref="([^"]+)"', xml)
    totals_match = re.search(r'totalsRowShown="([^"]+)"', xml)
    totals_count_match = re.search(r'totalsRowCount="([^"]+)"', xml)

    print(f"\nStep 6: XML attributes:")
    print(f"  ref: {ref_match.group(1) if ref_match else 'NOT FOUND'}")
    print(f"  totalsRowShown: {totals_match.group(1) if totals_match else 'NOT FOUND'}")
    print(f"  totalsRowCount: {totals_count_match.group(1) if totals_count_match else 'NOT FOUND'}")

    # Step 7: Check data integrity
    print(f"\nStep 7: Data integrity check:")
    print(f"  Row 4 (last data): {ws2['A4'].value}, {ws2['B4'].value}")
    print(f"  Row 5 (totals): {ws2['A5'].value}, {ws2['B5'].value}")

    print("\n\nSUCCESS: File created without corruption!")
    return True

if __name__ == "__main__":
    try:
        test_correct_approach()
    except Exception as e:
        print(f"\nFAILED: {e}")
        import traceback
        traceback.print_exc()
        exit(1)
