"""
FINAL FIX TEST: Apply totalsRowCount=1 to all tables with totals
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

def create_table_OLD_BROKEN(ws, df, table_name, sheet_name):
    """OLD BROKEN METHOD - missing totalsRowCount"""
    print(f"\nCreating {sheet_name} with OLD BROKEN method...")

    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Write data
    for row_idx in range(len(df)):
        for col_idx, col_name in enumerate(df.columns, 1):
            value = df.iloc[row_idx][col_name]
            if pd.isna(value):
                value = None
            elif hasattr(value, 'item'):
                value = value.item()
            ws.cell(row=row_idx + 2, column=col_idx, value=value)

    # Create table
    total_row = len(df) + 2
    tab_ref = f"A1:{get_column_letter(len(df.columns))}{total_row}"

    tab = Table(displayName=table_name, ref=tab_ref)
    tab.totalsRowShown = True
    # MISSING: tab.totalsRowCount = 1  <-- THIS IS THE BUG!
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium12",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    ws.add_table(tab)

    # Add totals formulas
    ws[f'A{total_row}'] = "Total"
    ws[f'B{total_row}'] = f"=SUBTOTAL(103,{table_name}[{df.columns[1]}])"

    print(f"  ✓ Table created: {table_name}")
    print(f"  ✓ Ref: {tab_ref}")
    print(f"  ✗ totalsRowCount: NOT SET (will cause corruption!)")

def create_table_NEW_FIXED(ws, df, table_name, sheet_name):
    """NEW FIXED METHOD - includes totalsRowCount=1"""
    print(f"\nCreating {sheet_name} with NEW FIXED method...")

    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Write data
    for row_idx in range(len(df)):
        for col_idx, col_name in enumerate(df.columns, 1):
            value = df.iloc[row_idx][col_name]
            if pd.isna(value):
                value = None
            elif hasattr(value, 'item'):
                value = value.item()
            ws.cell(row=row_idx + 2, column=col_idx, value=value)

    # Create table
    total_row = len(df) + 2
    tab_ref = f"A1:{get_column_letter(len(df.columns))}{total_row}"

    tab = Table(displayName=table_name, ref=tab_ref)
    tab.totalsRowShown = True
    tab.totalsRowCount = 1  # <-- THE FIX!
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium12",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    ws.add_table(tab)

    # Add totals formulas
    ws[f'A{total_row}'] = "Total"
    ws[f'B{total_row}'] = f"=SUBTOTAL(103,{table_name}[{df.columns[1]}])"

    print(f"  ✓ Table created: {table_name}")
    print(f"  ✓ Ref: {tab_ref}")
    print(f"  ✓ totalsRowCount: 1 (FIXED!)")

def main():
    print("="*70)
    print(" TESTING THE FIX FOR TABLE CORRUPTION")
    print("="*70)

    # Create test data similar to actual data
    df_mensagens = pd.DataFrame({
        'Data': ['10:00', '10:01', '10:02'],
        'Nome': ['Alice', 'Bob', 'Charlie'],
        'Mensagem': ['Hello', 'World', 'Test']
    })

    df_acessos = pd.DataFrame({
        'Nome': ['Alice', 'Bob'],
        'Tempo_Minutos': [90.5, 120.0]
    })

    df_inscritos = pd.DataFrame({
        'Nome': ['Alice', 'Bob', 'Charlie'],
        'Data de Cadastro': ['01/01/2024', '02/01/2024', '03/01/2024']
    })

    # Create workbook with OLD BROKEN tables
    print("\n" + "="*70)
    print(" PART 1: Creating file with BROKEN tables (no totalsRowCount)")
    print("="*70)

    wb_broken = Workbook()
    wb_broken.remove(wb_broken.active)

    ws1 = wb_broken.create_sheet("Mensagens")
    create_table_OLD_BROKEN(ws1, df_mensagens, "TABLE_MENSAGENS", "Mensagens")

    ws2 = wb_broken.create_sheet("Acessos")
    create_table_OLD_BROKEN(ws2, df_acessos, "TABLE_ACESSOS", "Acessos")

    ws3 = wb_broken.create_sheet("Inscritos")
    create_table_OLD_BROKEN(ws3, df_inscritos, "TABLE_INSCRITOS", "Inscritos")

    wb_broken.save("test_BROKEN_tables.xlsx")
    print("\n✗ Saved: test_BROKEN_tables.xlsx (will likely show corruption in Excel)")

    # Create workbook with NEW FIXED tables
    print("\n" + "="*70)
    print(" PART 2: Creating file with FIXED tables (with totalsRowCount=1)")
    print("="*70)

    wb_fixed = Workbook()
    wb_fixed.remove(wb_fixed.active)

    ws1 = wb_fixed.create_sheet("Mensagens")
    create_table_NEW_FIXED(ws1, df_mensagens, "TABLE_MENSAGENS", "Mensagens")

    ws2 = wb_fixed.create_sheet("Acessos")
    create_table_NEW_FIXED(ws2, df_acessos, "TABLE_ACESSOS", "Acessos")

    ws3 = wb_fixed.create_sheet("Inscritos")
    create_table_NEW_FIXED(ws3, df_inscritos, "TABLE_INSCRITOS", "Inscritos")

    wb_fixed.save("test_FIXED_tables.xlsx")
    print("\n✓ Saved: test_FIXED_tables.xlsx (should work perfectly in Excel)")

    print("\n" + "="*70)
    print(" INSTRUCTIONS:")
    print("="*70)
    print(" 1. Open test_BROKEN_tables.xlsx in Excel")
    print("    -> Should show corruption error for table2, table3, table4")
    print()
    print(" 2. Open test_FIXED_tables.xlsx in Excel")
    print("    -> Should open WITHOUT any errors!")
    print()
    print(" 3. If FIXED file works, apply the fix to excel_generator.py:")
    print("    -> Add 'tab.totalsRowCount = 1' after 'tab.totalsRowShown = True'")
    print("    -> In _create_table() method at line ~423")
    print("="*70)

if __name__ == "__main__":
    main()
