from openpyxl import load_workbook

wb = load_workbook('Temple PRSA  - estatísitcas - 2025-11-18.xlsx', data_only=False)
ws = wb['Mensagens']

print('Row 607:')
for col in 'ABCDEFGHIJ':
    print(f'  {col}: {ws[f"{col}607"].value}')

print('\nRow 608:')
for col in 'ABCDEFGHIJ':
    cell = ws[f"{col}608"]
    print(f'  {col}: value={cell.value}, type={type(cell.value).__name__}')
