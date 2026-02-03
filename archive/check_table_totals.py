#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Verificar formulas nas linhas totalizadoras do template
"""

import openpyxl

def check_totals():
    template_path = r"c:\Users\ferna\Desktop\App Estatisticas\Temple PRSA  - estatísitcas - 2025-11-18.xlsx"

    wb = openpyxl.load_workbook(template_path, data_only=False)

    print("=" * 80)
    print("VERIFICANDO LINHAS TOTALIZADORAS")
    print("=" * 80)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nPLANILHA: {sheet_name}")
        print("-" * 80)

        # Verificar última linha (normalmente é a totalizadora)
        max_row = ws.max_row
        print(f"Ultima linha: {max_row}")
        print(f"Conteudo da ultima linha:")

        for col_idx in range(1, min(10, ws.max_column + 1)):
            cell = ws.cell(row=max_row, column=col_idx)
            col_letter = openpyxl.utils.get_column_letter(col_idx)

            if cell.value:
                # Verificar se é fórmula
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    print(f"  {col_letter}{max_row}: FORMULA = {cell.value}")
                else:
                    value_str = str(cell.value)[:50]
                    print(f"  {col_letter}{max_row}: {value_str}")

if __name__ == "__main__":
    check_totals()
