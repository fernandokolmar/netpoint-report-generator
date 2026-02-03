import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Create test workbook
wb = Workbook()
ws = wb.active

# Test DataFrame with NaN and empty values
df = pd.DataFrame({
    'Name': ['Alice', 'Bob', 'Charlie'],
    'Value': [1, 2, 3],
    'Empty': [None, np.nan, '']
})

print("DataFrame:")
print(df)
print("\nDataFrame dtypes:")
print(df.dtypes)

# Write data using dataframe_to_rows
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        print(f"Row {r_idx}, Col {c_idx}: value={repr(value)}, cell.value={repr(cell.value)}")

# Create table
tab_ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 2}"
tab = Table(displayName="TestTable", ref=tab_ref)
tab.totalsRowShown = True
style = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)
tab.tableStyleInfo = style
ws.add_table(tab)

# Save and check
wb.save("test_empty_values.xlsx")
print(f"\nCreated table with ref: {tab_ref}")
print("File saved as test_empty_values.xlsx")
