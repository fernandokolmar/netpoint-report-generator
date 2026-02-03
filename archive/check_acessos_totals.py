#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Verificar formulas específicas da linha totalizadora de Acessos
"""

import openpyxl

def check_acessos():
    template_path = r"c:\Users\ferna\Desktop\App Estatisticas\Temple PRSA  - estatísitcas - 2025-11-18.xlsx"

    wb = openpyxl.load_workbook(template_path, data_only=False)
    ws = wb["Acessos"]

    max_row = ws.max_row
    print(f"Linha totalizadora (linha {max_row}):")
    print(f"  A{max_row}: {ws[f'A{max_row}'].value}")
    print(f"  B{max_row}: {ws[f'B{max_row}'].value}")
    print(f"  C{max_row}: {ws[f'C{max_row}'].value}")
    print(f"  D{max_row}: {ws[f'D{max_row}'].value}")
    print(f"  E{max_row}: {ws[f'E{max_row}'].value}")
    print(f"  F{max_row}: {ws[f'F{max_row}'].value}")
    print(f"  G{max_row}: {ws[f'G{max_row}'].value}")
    print(f"  H{max_row}: {ws[f'H{max_row}'].value}")

    # Verificar planilha Inscritos
    ws_insc = wb["Inscritos"]
    max_row_insc = ws_insc.max_row
    print(f"\nLinha totalizadora Inscritos (linha {max_row_insc}):")
    print(f"  A{max_row_insc}: {ws_insc[f'A{max_row_insc}'].value}")

    # Verificar últimas colunas
    for col_idx in range(1, ws_insc.max_column + 1):
        cell = ws_insc.cell(row=max_row_insc, column=col_idx)
        if cell.value:
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            print(f"  {col_letter}{max_row_insc}: {cell.value}")

if __name__ == "__main__":
    check_acessos()
