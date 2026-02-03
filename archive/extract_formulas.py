#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extrair formulas das linhas totalizadoras
"""

import openpyxl

def extract_formulas():
    template_path = r"c:\Users\ferna\Desktop\App Estatisticas\Temple PRSA  - estatísitcas - 2025-11-18.xlsx"
    output_file = r"c:\Users\ferna\Desktop\App Estatisticas\formulas_output.txt"

    wb = openpyxl.load_workbook(template_path, data_only=False)

    with open(output_file, 'w', encoding='utf-8') as f:
        # Acessos
        ws = wb["Acessos"]
        max_row = ws.max_row
        f.write(f"=== ACESSOS - Linha {max_row} ===\n")
        for col_idx in range(1, 9):
            cell = ws.cell(row=max_row, column=col_idx)
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            f.write(f"{col_letter}{max_row}: {repr(cell.value)}\n")

        # Inscritos
        ws_insc = wb["Inscritos"]
        max_row_insc = ws_insc.max_row
        f.write(f"\n=== INSCRITOS - Linha {max_row_insc} ===\n")
        for col_idx in range(1, ws_insc.max_column + 1):
            cell = ws_insc.cell(row=max_row_insc, column=col_idx)
            if cell.value:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                f.write(f"{col_letter}{max_row_insc}: {repr(cell.value)}\n")

        # Mensagens
        ws_msg = wb["Mensagens"]
        max_row_msg = ws_msg.max_row
        f.write(f"\n=== MENSAGENS - Linha {max_row_msg} ===\n")
        for col_idx in range(1, ws_msg.max_column + 1):
            cell = ws_msg.cell(row=max_row_msg, column=col_idx)
            if cell.value:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                f.write(f"{col_letter}{max_row_msg}: {repr(cell.value)}\n")

        # Retencao na live - células do resumo
        ws_ret = wb["Retenção na live"]
        f.write(f"\n=== RESUMO (Retenção na live) ===\n")
        for row in range(2, 8):
            f.write(f"Q{row}: {repr(ws_ret[f'Q{row}'].value)}\n")
            f.write(f"R{row}: {repr(ws_ret[f'R{row}'].value)}\n")

    print(f"Formulas extraídas para: {output_file}")

if __name__ == "__main__":
    extract_formulas()
